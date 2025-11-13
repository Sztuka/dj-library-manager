from __future__ import annotations
import argparse, csv, time, os, json
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

# --- Core importy (nasze moduły) ---
from djlib.config import (
    reconfigure, ensure_base_dirs, CONFIG_FILE,
    INBOX_DIR, READY_TO_PLAY_DIR, REVIEW_QUEUE_DIR, LOGS_DIR, CSV_PATH, AUDIO_EXTS
)
from djlib.csvdb import load_records, save_records
from djlib.tags import read_tags
from djlib.enrich import suggest_metadata, enrich_online_for_row
from djlib.genre import external_genre_votes, load_taxonomy_map, suggest_bucket_from_votes
from djlib.metadata.genre_resolver import resolve as resolve_genres
from djlib.classify import guess_bucket
from djlib.fingerprint import file_sha256, fingerprint_info
from djlib.filename import build_final_filename, extension_for
from djlib.mover import resolve_target_path, move_with_rename, utc_now_str
from djlib.buckets import is_valid_target
from djlib.placement import decide_bucket
try:
    from djlib.audio import check_env as audio_check_env
    from djlib.audio import analyze as audio_analyze
    from djlib.audio.cache import get_analysis
except Exception:
    # If audio backend is unavailable, fall back to None
    audio_check_env = None  # type: ignore
    audio_analyze = None  # type: ignore
    get_analysis = None  # type: ignore

# ML bucket assigner (optional)
try:
    from djlib.bucketing.simple_ml import SimpleMLBucketAssigner
except Exception:
    SimpleMLBucketAssigner = None  # type: ignore

# --- Pomocnicze ---
REPO_ROOT = Path(__file__).resolve().parents[1]

# ============ KOMENDY ============

def cmd_configure(_: argparse.Namespace) -> None:
    cfg, path = reconfigure()
    ensure_base_dirs()
    print(f"\n✅ Zapisano konfigurację do: {path}")
    print(f"   library_root: {cfg.library_root}")
    print(f"   inbox_dir:    {cfg.inbox_dir}\n")

def cmd_scan(_: argparse.Namespace) -> None:
    ensure_base_dirs()
    rows = load_records(CSV_PATH)
    known_hashes = {r.get("file_hash", "") for r in rows if r.get("file_hash")}
    known_fps    = {r.get("fingerprint", "") for r in rows if r.get("fingerprint")}

    # przygotuj plik statusu skanowania
    status_path = LOGS_DIR / "scan_status.json"
    def _write_status(data: Dict[str, Any]) -> None:
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            with status_path.open("w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
        except Exception:
            pass

    all_files = [p for p in INBOX_DIR.glob("**/*") if p.is_file() and p.suffix.lower() in AUDIO_EXTS]
    total = len(all_files)
    processed = 0
    added = 0
    errors = 0
    missing_fpcalc = False
    _write_status({
        "state": "running",
        "total": total,
        "processed": 0,
        "added": 0,
        "errors": 0,
        "last_file": "",
    })

    new_rows = []
    for p in all_files:
        fhash = file_sha256(p)
        if fhash in known_hashes:
            processed += 1
            _write_status({"state": "running", "total": total, "processed": processed, "added": added, "errors": errors, "last_file": str(p)})
            continue

        tags = read_tags(p)
        try:
            dur, fp = fingerprint_info(p)
        except Exception as e:
            fp = ""
            dur = 0
            errors += 1
            if "fpcalc" in str(e).lower():
                missing_fpcalc = True

        is_dup = "true" if (fp and fp in known_fps) else "false"

        ai_bucket, ai_comment = guess_bucket(
            tags["artist"], tags["title"], tags["bpm"], tags["genre"], tags["comment"]
        )

        # Proponowane metadane (preferuj online w przyszłości; teraz filename+fallback)
        sugg = suggest_metadata(p, tags)
        # jeśli nie mamy duration z online, wstaw lokalny czas z fingerprintu
        if (sugg.get("duration_suggest") or "").strip() == "" and dur:
            mm = dur // 60
            ss = dur % 60
            sugg["duration_suggest"] = f"{mm}:{ss:02d}"

        track_id = f"{fhash[:12]}_{int(time.time())}"
        rec = {
            "track_id": track_id,
            "file_path": str(p),
            # Wypełnij główne pola z metatagów jeśli dostępne (ułatwia eksport / filtrowanie)
            "artist": (tags.get("artist") or "").strip(),
            "title": (tags.get("title") or "").strip(),
            "version_info": (tags.get("version_info") or "").strip(),
            "bpm": tags["bpm"],
            "key_camelot": tags["key_camelot"],
            "energy_hint": tags["energy_hint"],
            "file_hash": fhash,
            "fingerprint": fp,
            "is_duplicate": is_dup,
            "ai_guess_bucket": ai_bucket,
            "ai_guess_comment": ai_comment,
            "target_subfolder": "",
            "must_play": "",
            "occasion_tags": "",
            "notes": "",
            "final_filename": "",
            "final_path": "",
            "added_date": "",
            # propozycje do weryfikacji
            **sugg,
            "review_status": "pending",
        }
        new_rows.append(rec)
        known_hashes.add(fhash)
        if fp:
            known_fps.add(fp)
        added += 1
        processed += 1
        _write_status({
            "state": "running",
            "total": total,
            "processed": processed,
            "added": added,
            "errors": errors,
            "last_file": str(p),
            "missing_fpcalc": missing_fpcalc,
        })

    if new_rows:
        rows.extend(new_rows)
        save_records(CSV_PATH, rows)
        print(f"Zeskanowano {len(new_rows)} plików. Zapisano {CSV_PATH}.")
    else:
        print("Brak nowych plików do dodania.")

    _write_status({
        "state": "done",
        "total": total,
        "processed": processed,
        "added": added,
        "errors": errors,
        "csv_rows": len(load_records(CSV_PATH)),
        "missing_fpcalc": missing_fpcalc,
    })

def _load_rules(path: Path) -> Dict[str, Any]:
    import yaml
    if not path.exists():
        return {"rules": [], "fallbacks": {}}
    with path.open("r", encoding="utf-8") as f:
        return (yaml.safe_load(f) or {"rules": [], "fallbacks": {}})

def _decide_for_row(row: Dict[str, str], rules: Dict[str, Any]) -> str:
    artist = (row.get("artist") or "").lower()
    title  = (row.get("title") or "").lower()
    genre  = (row.get("genre") or "").lower()
    comm   = (row.get("ai_guess_comment") or row.get("comment") or "").lower()
    haystack = " ".join([artist, title, genre, comm])

    for rule in rules.get("rules", []):
        words = [w.lower() for w in rule.get("contains", [])]
        if any(w in haystack for w in words):
            return rule.get("target", "")

    fb = rules.get("fallbacks", {})
    guess = row.get("ai_guess_bucket") or ""
    if guess in fb:
        return fb[guess]
    return fb.get("default", "REVIEW QUEUE/UNDECIDED")

def cmd_auto_decide(args: argparse.Namespace) -> None:
    rules_path = Path(args.rules or (REPO_ROOT / "rules.yml"))
    rules = _load_rules(rules_path)
    rows = load_records(CSV_PATH)
    updated = 0

    for r in rows:
        if args.only_empty and (r.get("target_subfolder") or "").strip():
            continue
        proposal = _decide_for_row(r, rules)
        if is_valid_target(proposal):
            r["target_subfolder"] = proposal
            updated += 1

    if updated:
        save_records(CSV_PATH, rows)
        print(f"Zaktualizowano {updated} wierszy.")
    else:
        print("Brak zmian.")

def cmd_auto_decide_smart(_: argparse.Namespace) -> None:
    """Lepsze auto-decide: używa heurystyk z djlib.placement z progami ufności.
    ≥0.85: ustaw docelowy kubełek; 0.65..0.85: tylko sugestia (ai_guess_*)."""
    HARDCOMMIT_CONF = 0.85
    SUGGEST_CONF = 0.65
    rows = load_records(CSV_PATH)
    set_cnt = sug_cnt = 0
    for r in rows:
        if r.get("target_subfolder"):
            continue
        tgt, conf, reason = decide_bucket(r)
        if not tgt:
            continue
        if conf >= HARDCOMMIT_CONF:
            r["target_subfolder"] = f"READY TO PLAY/{tgt}"
            r["ai_guess_bucket"] = ""
            r["ai_guess_comment"] = f"rule:{reason}; conf={conf:.2f}"
            set_cnt += 1
        elif conf >= SUGGEST_CONF:
            r["ai_guess_bucket"]  = f"READY TO PLAY/{tgt}"
            r["ai_guess_comment"] = f"rule:{reason}; conf={conf:.2f}"
            sug_cnt += 1
    if set_cnt or sug_cnt:
        save_records(CSV_PATH, rows)
    print(f"✅ Auto-decide (smart): set={set_cnt}, suggested={sug_cnt}")

def cmd_enrich_online(_: argparse.Namespace) -> None:
    """Wzbogaca metadane (suggest_*) dla pozycji pending korzystając z MusicBrainz/AcoustID/Last.fm/Spotify.
    Prowadzi status w LOGS/enrich_status.json, aby UI mogło pokazywać postęp.
    Nie nadpisuje już zaakceptowanych. Nie zmienia BPM/Key.
    """
    rows = load_records(CSV_PATH)
    # Jeśli podano --force-genres w args, będziemy nadpisywać źródłowe kolumny even if present
    force_genres = bool(getattr(_, "force_genres", False))  # '_' is args Namespace
    todo = [r for r in rows if (r.get("review_status") or "").lower() != "accepted"]
    total = len(todo)
    processed = 0
    changed = 0
    mb_set = 0
    lfm_set = 0
    sp_set = 0

    # Check API credentials presence for diagnostics
    try:
        from djlib.config import get_lastfm_api_key, get_spotify_credentials
        _lfm_key_present = bool(get_lastfm_api_key())
        _sp_cid, _sp_sec = get_spotify_credentials()
        _sp_present = bool(_sp_cid and _sp_sec)
    except Exception:
        _lfm_key_present = False
        _sp_present = False

    # status plik
    status_path = LOGS_DIR / "enrich_status.json"
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    def _now_iso() -> str:
        return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    # Struktura statusu (rozszerzona zgodnie z ARCHITECTURE.md)
    status_doc = {
        "started_at": _now_iso(),
        "completed_at": "",
        "rows_total": total,
        "rows_processed": 0,
        "updated": 0,
        "state": "running",
        "last_file": "",
        "soundcloud": {
            "client_id_status": "unknown",
            "decision": "pending",  # active | skipped | aborted
            "prompt_shown": False,
            "attempted_requests": 0,
            "timestamp": _now_iso(),
        },
        "sources_counts": {
            "musicbrainz": 0,
            "lastfm": 0,
            "spotify": 0,
            "soundcloud": 0,
        }
    }

    def _flush_status() -> None:
        try:
            with status_path.open("w", encoding="utf-8") as f:
                json.dump(status_doc, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    _flush_status()

    # SoundCloud client id health (informative, does not block)
    sc_health_msg = ""
    try:
        from djlib.metadata.soundcloud import client_id_health
        h = client_id_health()
        if h:
            status_doc["soundcloud"]["client_id_status"] = h.get("status", "unknown")
        sc_health_msg = f"soundcloud_client_id_status={h.get('status')}" if h else ""
        if h and h.get("status") == "ok":
            print("✅ SoundCloud client_id OK")
            if not getattr(_, "skip_soundcloud", False):
                status_doc["soundcloud"]["decision"] = "active"
        elif h and h.get("status") in {"invalid", "error"}:
            print(f"⚠ SoundCloud client_id: {h.get('message')}")
            if getattr(_, "skip_soundcloud", False):
                status_doc["soundcloud"]["decision"] = "skipped"
            else:
                status_doc["soundcloud"]["prompt_shown"] = True
                _flush_status()
                try:
                    choice = input("Kontynuować bez SoundCloud? [Y/n]: ").strip().lower()
                except Exception:
                    choice = "y"
                if choice in {"n", "no"}:
                    print("Przerwano na prośbę użytkownika (SoundCloud invalid).")
                    status_doc["soundcloud"]["decision"] = "aborted"
                    status_doc["state"] = "done"
                    status_doc["completed_at"] = _now_iso()
                    _flush_status()
                    return
                else:
                    print("→ Pomiń SoundCloud w tym przebiegu.")
                    setattr(_, "skip_soundcloud", True)
                    status_doc["soundcloud"]["decision"] = "skipped"
        elif h and h.get("status") == "missing":
            if getattr(_, "skip_soundcloud", False):
                status_doc["soundcloud"]["decision"] = "skipped"
            else:
                print("ℹ Brak SoundCloud client_id (można ustawić DJLIB_SOUNDCLOUD_CLIENT_ID).")
                status_doc["soundcloud"]["decision"] = "skipped"  # treat missing as skipped
    except Exception:
        pass
    _flush_status()

    # przygotuj mapowanie tagów → bucket
    tag_map = load_taxonomy_map()

    for r in rows:
        if (r.get("review_status") or "").lower() == "accepted":
            continue
        p = Path(r.get("file_path",""))
        online = enrich_online_for_row(p, r)
        if not online:
            processed += 1
            status_doc["rows_processed"] = processed
            status_doc["updated"] = changed
            status_doc["last_file"] = str(p)
            _flush_status()
            continue
        # reguła nadpisywania:
        # - zawsze nadpisuj, jeśli źródłem jest AcoustID (najwyższy priorytet)
        # - w innym przypadku: wypełnij jeśli puste LUB nadpisz fallback (filename|tags_fallback)
        current_source = (r.get("meta_source") or "").strip().lower()
        online_source = (online.get("meta_source") or "").strip().lower()
        acoustid_wins = online_source.startswith("acoustid")
        allow_override = acoustid_wins or (
            current_source in {"filename|tags_fallback", "filename,tags_fallback", "tags_fallback"}
        ) or not r.get("genre_suggest")  # nadpisz jeśli genre pusty
        any_change = False
        for k, v in online.items():
            if k in {"artist_suggest","title_suggest","version_suggest","genre_suggest","album_suggest","year_suggest","duration_suggest"}:
                cur = (r.get(k) or "").strip()
                if (not cur and v) or (allow_override and v and cur != v):
                    r[k] = v
                    any_change = True
        # ustaw meta_source jeśli zrobiliśmy jakąkolwiek aktualizację i online podał źródło
        if any_change and (online.get("meta_source") or "").strip():
            r["meta_source"] = online["meta_source"]
        
        # Zawsze spróbuj wzbogacić gatunki używając wszystkich źródeł (MB + Last.fm + Spotify)
        try:
            a = (r.get("artist_suggest") or r.get("artist") or "").strip()
            t = (r.get("title_suggest") or r.get("title") or "").strip()
            dur_s = None
            if r.get("duration_suggest"):
                try:
                    dur_parts = r["duration_suggest"].split(":")
                    if len(dur_parts) == 2:
                        dur_s = int(dur_parts[0]) * 60 + int(dur_parts[1])
                except Exception:
                    pass
            
            from djlib.metadata.genre_resolver import resolve as resolve_genres
            genre_res = resolve_genres(a, t, duration_s=dur_s, disable_soundcloud=bool(getattr(_, "skip_soundcloud", False)))
            if genre_res and genre_res.confidence >= 0.03:  # lower threshold for missing genres
                # Ustaw 3 gatunki: main + subs
                genres = [genre_res.main] + genre_res.subs[:2]  # max 3 total
                genre_str = ", ".join(genres)
                current_genre = (r.get("genre_suggest") or "").strip()
                if not current_genre or genre_res.confidence > 0.08:  # override existing only if significantly better
                    r["genre_suggest"] = genre_str
                    any_change = True
                    # Update meta_source to reflect all sources used
                    sources = [src for src, _, _ in genre_res.breakdown]
                    if sources:
                        r["meta_source"] = f"{r.get('meta_source', '')}+genres({','.join(sources)})".strip("+")

                # Zapisz surowe listy tagów per źródło do dodatkowych kolumn
                try:
                    src_map = {src: local for (src, _, local) in genre_res.breakdown}
                    def _top_k(d, k=5):
                        return ", ".join([kv[0] for kv in sorted(d.items(), key=lambda kv: kv[1], reverse=True)[:k]])
                    if src_map.get("musicbrainz") and (force_genres or not (r.get("genres_musicbrainz") or "")):
                        r["genres_musicbrainz"] = _top_k(src_map["musicbrainz"])  # type: ignore[index]
                        any_change = True
                        mb_set += 1
                    if src_map.get("lastfm") and (force_genres or not (r.get("genres_lastfm") or "")):
                        r["genres_lastfm"] = _top_k(src_map["lastfm"])  # type: ignore[index]
                        any_change = True
                        lfm_set += 1
                    if src_map.get("spotify") and (force_genres or not (r.get("genres_spotify") or "")):
                        r["genres_spotify"] = _top_k(src_map["spotify"])  # type: ignore[index]
                        any_change = True
                        sp_set += 1
                    if src_map.get("soundcloud") and (force_genres or not (r.get("genres_soundcloud") or "")):
                        r["genres_soundcloud"] = _top_k(src_map["soundcloud"])  # type: ignore[index]
                        any_change = True
                except Exception:
                    pass
        except Exception as e:
            # Debug: print exception for troubleshooting
            print(f"Genre resolution failed for {a} - {t}: {e}")
            pass

        # Popularność z Last.fm (playcount/listeners) — pomoc dla singalong/party dance/decades
        try:
            a = (r.get("artist_suggest") or r.get("artist") or "").strip()
            t = (r.get("title_suggest") or r.get("title") or "").strip()
            if a and t:
                from djlib.metadata.lastfm import track_info as lf_track_info
                info = lf_track_info(a, t) or {}
                if info:
                    # Zapisz pola popularności, nie nadpisuj istniejących >0
                    if info.get("playcount") and int(info.get("playcount", 0)) > int(r.get("pop_playcount", 0) or 0):
                        r["pop_playcount"] = str(info["playcount"])  # zapis w CSV jako string
                    if info.get("listeners") and int(info.get("listeners", 0)) > int(r.get("pop_listeners", 0) or 0):
                        r["pop_listeners"] = str(info["listeners"])  # zapis w CSV jako string
        except Exception:
            pass

        # Zaproponuj kubełek na podstawie gatunków
        try:
            genre_str = (r.get("genre_suggest") or "").strip()
            if genre_str and tag_map:
                # Parse genres back to individual tags for voting
                genre_tags = [g.strip() for g in genre_str.split(",") if g.strip()]
                votes = {tag: 1.0 for tag in genre_tags}  # equal weight for each genre
                bucket, conf, breakdown = suggest_bucket_from_votes(votes, tag_map)
                if bucket and conf >= 0.65:
                    r["ai_guess_bucket"]  = f"READY TO PLAY/{bucket}"
                    # zbuduj krótki komentarz z top tagów
                    top_tags = [tag for tag, _, mapped in breakdown if mapped][:3]
                    tags_str = ", ".join(top_tags) if top_tags else genre_str.split(",")[0]
                    r["ai_guess_comment"] = f"genres; conf={conf:.2f}; tags: {tags_str}"
                    any_change = True
        except Exception:
            pass

        if any_change:
            changed += 1
        # Auto-fill artist/title if still empty and we now have suggest values (quality-of-life)
        if not (r.get("artist") or "").strip() and (r.get("artist_suggest") or "").strip():
            r["artist"] = r["artist_suggest"]
        if not (r.get("title") or "").strip() and (r.get("title_suggest") or "").strip():
            r["title"] = r["title_suggest"]
        processed += 1
        status_doc["rows_processed"] = processed
        status_doc["updated"] = changed
        status_doc["last_file"] = str(p)
        _flush_status()
    if changed:
        save_records(CSV_PATH, rows)
    # Oblicz źródła użycia na podstawie wypełnionych kolumn per-source
    mb_cnt = lfm_cnt = sp_cnt = sc_cnt = 0
    for r in rows:
        if r.get("genres_musicbrainz"):
            mb_cnt += 1
        if r.get("genres_lastfm"):
            lfm_cnt += 1
        if r.get("genres_spotify"):
            sp_cnt += 1
        if r.get("genres_soundcloud"):
            sc_cnt += 1
    status_doc["sources_counts"] = {
        "musicbrainz": mb_cnt,
        "lastfm": lfm_cnt,
        "spotify": sp_cnt,
        "soundcloud": sc_cnt,
    }
    # Uzupełnij attempted_requests z modułu SoundCloud
    try:
        from djlib.metadata.soundcloud import soundcloud_request_count
        status_doc["soundcloud"]["attempted_requests"] = soundcloud_request_count()
    except Exception:
        pass
    status_doc["rows_processed"] = processed
    status_doc["updated"] = changed
    status_doc["state"] = "done"
    status_doc["completed_at"] = _now_iso()
    _flush_status()
    print(f"🔎 Enrich online: updated={changed}")
    # Short diagnostics
    if total:
        print(f"   → genres set — MB:{mb_set}, LFM:{lfm_set}, SP:{sp_set}")
    if not _lfm_key_present:
        print("   ⚠ Brak LASTFM_API_KEY (DJLIB_LASTFM_API_KEY) — kolumna genres_lastfm może pozostać pusta.")
    if not _sp_present:
        print("   ⚠ Brak SPOTIFY client credentials (DJLIB_SPOTIFY_CLIENT_ID/SECRET) — kolumna genres_spotify może pozostać pusta.")
    if sc_health_msg:
        print(f"   ℹ {sc_health_msg}")

def cmd_fix_fingerprints(_: argparse.Namespace) -> None:
    """Uzupełnij brakujące fingerprinty w istniejącym CSV.
    Dla każdego wiersza bez fingerprintu spróbuj wyliczyć go z pliku (preferuj final_path, potem file_path).
    Aktualizuj też duration_suggest jeśli puste.
    Pisz postęp do LOGS/fingerprint_status.json, aby UI mogło pokazywać pasek.
    """
    from djlib.config import LOGS_DIR
    rows = load_records(CSV_PATH)
    targets = []
    for r in rows:
        fp = (r.get("fingerprint") or "").strip()
        if fp:
            continue
        # preferuj final_path jeśli istnieje, w przeciwnym razie file_path
        p = None
        fp1 = r.get("final_path") or ""
        fp2 = r.get("file_path") or ""
        f1 = Path(fp1) if fp1 else None
        f2 = Path(fp2) if fp2 else None
        if f1 and f1.exists():
            p = f1
        elif f2 and f2.exists():
            p = f2
        if p is not None:
            targets.append((r, p))

    total = len(targets)
    processed = 0
    updated = 0
    errors = 0

    status_path = LOGS_DIR / "fingerprint_status.json"

    def _write_status(state: str, last_file: str = "") -> None:
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            with status_path.open("w", encoding="utf-8") as f:
                json.dump({
                    "state": state,
                    "total": total,
                    "processed": processed,
                    "updated": updated,
                    "errors": errors,
                    "last_file": last_file,
                }, f, ensure_ascii=False)
        except Exception:
            pass

    _write_status("running", "")

    for r, p in targets:
        try:
            dur, fp = fingerprint_info(p)
            if fp:
                r["fingerprint"] = fp
                # uzupełnij duration_suggest jeśli brak
                ds = (r.get("duration_suggest") or "").strip()
                if not ds and dur:
                    mm, ss = divmod(int(dur), 60)
                    r["duration_suggest"] = f"{mm}:{ss:02d}"
                updated += 1
        except Exception:
            errors += 1
        processed += 1
        _write_status("running", str(p))

    if updated:
        save_records(CSV_PATH, rows)
    _write_status("done", "")
    print(f"🧩 Fix fingerprints: updated={updated}, errors={errors}")

def cmd_apply(args: argparse.Namespace) -> None:
    rows = load_records(CSV_PATH)
    changed = False

    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d-%H%M%S")
    log_path = LOGS_DIR / f"moves-{stamp}.csv"
    log_rows = []

    for r in rows:
        target = (r.get("target_subfolder") or "").strip()
        if not target or target.upper() == "REJECT":
            continue
        if r.get("final_path"):
            continue

        src = Path(r["file_path"])
        if not src.exists():
            print(f"[WARN] Nie znaleziono pliku: {src}")
            continue

        dest_dir = resolve_target_path(target)
        if dest_dir is None:
            continue

        final_name = build_final_filename(
            r.get("artist_suggest") or r.get("artist", ""), 
            r.get("title_suggest") or r.get("title", ""),
            r.get("version_suggest") or r.get("version_info", ""), 
            r.get("key_camelot", ""),
            r.get("bpm", ""), 
            extension_for(src)
        )

        dest_path = dest_dir / final_name
        print(f"{'DRY-RUN ' if args.dry_run else ''}MOVE: {src} -> {dest_path}")

        if args.dry_run:
            continue

        dest_real = move_with_rename(src, dest_dir, final_name)
        r["final_filename"] = final_name
        r["final_path"] = str(dest_real)
        r["added_date"] = utc_now_str()
        changed = True
        log_rows.append([str(src), str(dest_real), r.get("track_id","")])

    if not args.dry_run and log_rows:
        with log_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["src_before", "dest_after", "track_id"])
            w.writerows(log_rows)
        print(f"Zapisano log: {log_path}")

    if changed and not args.dry_run:
        save_records(CSV_PATH, rows)
        print("Przeniesiono i zaktualizowano CSV.")
    elif not changed and not args.dry_run:
        print("Brak pozycji do przeniesienia.")

def scan_command() -> None:
    """Funkcja wywołująca skanowanie (używana przez webapp i inne moduły)."""
    args = argparse.Namespace()
    cmd_scan(args)

def cmd_undo(_: argparse.Namespace) -> None:
    logs = sorted(LOGS_DIR.glob("moves-*.csv"))
    if not logs:
        print("Brak logów do cofnięcia.")
        return
    log = logs[-1]
    print(f"Cofam ruchy z: {log.name}")

    rows = list(csv.DictReader(log.open("r", encoding="utf-8")))
    reverted = 0
    for r in rows:
        src_before = Path(r["src_before"])
        dest_after = Path(r["dest_after"])
        if dest_after.exists():
            dest_after.rename(src_before)
            reverted += 1
        else:
            print(f"[WARN] Brak pliku do cofnięcia: {dest_after}")
    print(f"Cofnięto {reverted} ruchów.")

def cmd_dupes(_: argparse.Namespace) -> None:
    groups: dict[str, list[dict[str,str]]] = {}
    rows = load_records(CSV_PATH)
    for r in rows:
        fp = (r.get("fingerprint") or "").strip()
        if not fp:
            continue
        groups.setdefault(fp, []).append(r)

    out = LOGS_DIR / "dupes.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["group_fingerprint", "track_id", "artist", "title", "file_path", "final_path", "file_hash"])
        for fp, items in groups.items():
            if len(items) <= 1:
                continue
            for r in items:
                w.writerow([fp, r.get("track_id",""), r.get("artist",""), r.get("title",""),
                            r.get("file_path",""), r.get("final_path",""), r.get("file_hash","")])
    print(f"Zapisano raport duplikatów: {out}")

def cmd_sync_audio_metrics(args: argparse.Namespace) -> None:
    """Zsynchronizuj metryki (BPM/Key/Energy) z cache SQLite do głównego CSV.
    Domyślnie uzupełnia tylko puste pola; użyj --force aby nadpisać istniejące.
    Opcjonalnie zapisuje metadane do plików audio jeśli --write-tags.
    """
    if get_analysis is None:
        print("Audio cache backend niedostępny.")
        return
    rows = load_records(CSV_PATH)
    if not rows:
        print("Brak wierszy w CSV.")
        return
    force = bool(getattr(args, "force", False))
    write_tags_flag = bool(getattr(args, "write_tags", False))
    updated = 0
    tags_written = 0
    for r in rows:
        audio_id = (r.get("file_hash") or "").strip()
        if not audio_id:
            # spróbuj policzyć hash jeśli plik istnieje
            try:
                p = Path(r.get("file_path") or "")
                if p.exists():
                    from djlib.audio.cache import compute_audio_id as _cmp
                    audio_id = _cmp(p)
                else:
                    continue
            except Exception:
                continue
        a = get_analysis(audio_id)
        if not a:
            continue
        # przygotuj wartości
        bpm = a.get("bpm")
        key = a.get("key_camelot")
        energy = a.get("energy")
        # uzupełniaj tylko puste chyba że --force
        def _should_set(cur: str) -> bool:
            return force or (not (cur or "").strip())
        changed = False
        if bpm is not None and _should_set(r.get("bpm", "")):
            try:
                r["bpm"] = f"{float(bpm):.2f}".rstrip("0").rstrip(".")
            except Exception:
                r["bpm"] = str(bpm)
            changed = True
        if key and _should_set(r.get("key_camelot", "")):
            r["key_camelot"] = str(key)
            changed = True
        if energy is not None and _should_set(r.get("energy_hint", "")):
            # energy jako 0..1 → wpisz procentowo (0..100) lub float; wybierz prosty procent
            try:
                r["energy_hint"] = f"{round(float(energy)*100)}"
            except Exception:
                r["energy_hint"] = str(energy)
            changed = True
        if changed:
            updated += 1
        
        # Zapisz metadane do pliku jeśli --write-tags
        if write_tags_flag and (bpm or key):
            try:
                p = Path(r.get("file_path") or "")
                if p.exists():
                    from djlib.tags import write_tags
                    tag_updates = {}
                    if bpm is not None:
                        tag_updates["bpm"] = str(bpm)
                    if key:
                        tag_updates["key_camelot"] = str(key)
                    write_tags(p, tag_updates)
                    tags_written += 1
            except Exception as e:
                print(f"[WARN] Nie udało się zapisać tagów do {p}: {e}")
    
    if updated:
        save_records(CSV_PATH, rows)
    print(f"🔄 Sync audio metrics: updated={updated}, tags_written={tags_written}")

def cmd_genres_resolve(args: argparse.Namespace) -> None:
    artist = (getattr(args, "artist", None) or "").strip()
    title = (getattr(args, "title", None) or "").strip()
    dur = getattr(args, "duration", None)
    res = resolve_genres(artist, title, duration_s=dur)
    if not res:
        print("Brak wyników z zewnętrznych źródeł (MB/LFM/Spotify).")
        return
    print(f"Main: {res.main}")
    if res.subs:
        print(f"Subs: {', '.join(res.subs)}")
    print(f"Confidence: {res.confidence:.2f}")
    print("Breakdown:")
    for src, _, local in res.breakdown:
        parts = ", ".join(f"{k}:{v:.2f}" for k, v in sorted(local.items(), key=lambda kv: kv[1], reverse=True)[:5])
        print(f"  - {src}: {parts}")

def cmd_detect_taxonomy(_: argparse.Namespace) -> None:
    """Wykrywa istniejącą strukturę folderów i zapisuje jako taxonomy.local.yml."""
    from djlib.taxonomy import detect_taxonomy_from_fs, save_taxonomy, load_taxonomy
    from djlib.config import LIB_ROOT

    # Załaduj istniejącą taksonomię
    existing = load_taxonomy()
    existing_ready = set(existing["ready_buckets"])
    existing_review = set(existing["review_buckets"])

    # Wykryj nową z filesystem
    detected = detect_taxonomy_from_fs(LIB_ROOT)
    detected_ready = set(detected["ready_buckets"])
    detected_review = set(detected["review_buckets"])

    # Merge: dodaj nowe wykryte, zachowaj istniejące
    merged_ready = existing_ready | detected_ready
    merged_review = existing_review | detected_review

    merged = {
        "ready_buckets": sorted(merged_ready),
        "review_buckets": sorted(merged_review),
    }

    save_taxonomy(merged)
    print(f"Zaktualizowano taksonomię: {len(merged_ready)} ready buckets, {len(merged_review)} review buckets")
    if merged_ready:
        print("Ready buckets:", ", ".join(merged_ready))
    if merged_review:
        print("Review buckets:", ", ".join(merged_review))


def cmd_taxonomy_backup(_: argparse.Namespace) -> None:
    """Zrób snapshot taksonomii na podstawie realnej struktury folderów (LIB_ROOT) i zapisz do backupów.

    Nie modyfikuje istniejącego taxonomy.local.yml. Tworzy:
    - taxonomy.local.yml.backup (nadpisywalny snapshot)
    - taxonomy.local.<timestamp>.yml (archiwalny snapshot)
    """
    from djlib.taxonomy import detect_taxonomy_from_fs
    from djlib.config import LIB_ROOT
    import yaml as _yaml

    lib_root = LIB_ROOT
    data = detect_taxonomy_from_fs(lib_root)
    stamp = time.strftime("%Y%m%d-%H%M%S")
    backup_path = REPO_ROOT / "taxonomy.local.yml.backup"
    archive_path = REPO_ROOT / f"taxonomy.local.{stamp}.yml"
    try:
        with backup_path.open("w", encoding="utf-8") as f:
            _yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)
        with archive_path.open("w", encoding="utf-8") as f:
            _yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)
        print(f"📦 Snapshot zapisany: {backup_path} oraz {archive_path}")
    except Exception as e:
        print(f"[ERR] Nie udało się zapisać backupu taksonomii: {e}")

def cmd_analyze_audio(args: argparse.Namespace) -> None:
    """Analiza audio (BPM/Key/Energy) dla INBOX lub wskazanego pliku/katalogu.
    Wyniki zapisywane są do cache SQLite (LOGS/audio_analysis.sqlite).
    """
    # Obsłuż --check-env
    if getattr(args, "check_env", False):
        if audio_check_env is None:
            print("Essentia backend niedostępny (brak modułu).")
            return
        info = audio_check_env()
        print(json.dumps(info, ensure_ascii=False, indent=2))
        return

    if audio_analyze is None:
        print("Audio backend niedostępny. Zainstaluj Essentia lub uruchom z --check-env, aby sprawdzić środowisko.")
        return

    # Zbierz pliki do analizy
    targets = []
    base = Path(getattr(args, "path", "") or INBOX_DIR)
    if base.is_file():
        targets = [base]
    else:
        base = base if base.exists() else INBOX_DIR
        targets = [p for p in base.glob("**/*") if p.is_file() and p.suffix.lower() in AUDIO_EXTS]

    total = len(targets)
    print(f"DEBUG: base={base}, total_targets={total}")  # DEBUG
    processed = 0
    updated = 0
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    status_path = LOGS_DIR / "audio_status.json"

    def _write_status(state: str, last_file: str = "", last_error: str = "") -> None:
        try:
            with status_path.open("w", encoding="utf-8") as f:
                json.dump({
                    "state": state,
                    "total": total,
                    "processed": processed,
                    "updated": updated,
                    "last_file": last_file,
                    "error": last_error,
                }, f, ensure_ascii=False)
        except Exception:
            pass

    # Parsuj zakres BPM
    lo, hi = 80, 180
    tb = getattr(args, "target_bpm", None)
    if tb and ":" in tb:
        try:
            lo, hi = [int(x) for x in tb.split(":", 1)]
        except Exception:
            pass

    _write_status("running", "")
    for p in targets:
        print(f"DEBUG: processing {p}")  # DEBUG
        try:
            res = audio_analyze(p, target_bpm_range=(lo, hi), recompute=bool(args.recompute), config={"target_bpm": [lo, hi]})
            # Jeśli analyze dokonało upsert do cache, liczymy jako updated
            if res:
                updated += 1
            _write_status("running", str(p))
        except Exception as e:
            print(f"DEBUG: exception {e}")  # DEBUG
            _write_status("running", str(p), str(e))
        processed += 1

    _write_status("done", "")
    print(f"🎧 Analyze-audio: files={total}, analyzed={updated}")

def _normalize_ml_features_from_analysis(res: Dict[str, Any]) -> Dict[str, Any]:
    """Map fields from analyze() payload to ML feature dict expected by SimpleMLBucketAssigner."""
    # Keep only scalar-like entries; ignore meta keys
    feat: Dict[str, Any] = {}
    for k, v in res.items():
        if k in {"algo_version", "config_hash", "analyzed_at", "source", "audio_id"}:
            continue
        feat[k] = v
    # Aliases
    if "bpm_detected" not in feat and "bpm" in res:
        feat["bpm_detected"] = res.get("bpm")
    if "energy_score" not in feat and "energy" in res:
        feat["energy_score"] = res.get("energy")
    return feat

def cmd_ml_predict(args: argparse.Namespace) -> None:
    if SimpleMLBucketAssigner is None:
        print("ML assigner not available (scikit-learn missing or import failure).")
        return
    if audio_analyze is None:
        print("Audio analyze backend niedostępny — zainstaluj Essentia lub użyj analyze-audio --check-env.")
        return

    model_path = Path(getattr(args, "model", "") or (REPO_ROOT / "models" / "fma_trained_model_balanced.pkl"))
    if not model_path.exists():
        print(f"Brak modelu ML pod: {model_path}")
        return
    assigner = SimpleMLBucketAssigner(model_path)

    base = Path(getattr(args, "path", "") or INBOX_DIR)
    targets = [base] if base.is_file() else [p for p in base.glob("**/*") if p.is_file() and p.suffix.lower() in AUDIO_EXTS]
    total = len(targets)
    print(f"ML Predict: model={model_path}, files={total}")

    rows = load_records(CSV_PATH)
    by_path = {r.get("file_path"): r for r in rows}
    hard_t = float(getattr(args, "hard_threshold", 0.85))
    suggest_t = float(getattr(args, "suggest_threshold", 0.65))
    min_conf = float(getattr(args, "min_confidence", 0.40))
    set_cnt = sug_cnt = 0

    log_path = LOGS_DIR / "ml_predictions.csv"
    if not log_path.exists():
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            with log_path.open("w", encoding="utf-8") as f:
                f.write("file_path,bucket,confidence\n")
        except Exception:
            pass

    for p in targets:
        try:
            res = audio_analyze(p, recompute=bool(getattr(args, "recompute", False))) or {}
            feat = _normalize_ml_features_from_analysis(res)
            bucket, conf = assigner.predict(feat)
            # Log
            try:
                with log_path.open("a", encoding="utf-8") as f:
                    f.write(f"{p},{bucket},{conf:.4f}\n")
            except Exception:
                pass
            print(f"{p.name}: {bucket} (conf={conf:.2f})")

            # Optionally write to CSV
            if getattr(args, "set_target", False) or getattr(args, "suggest", False):
                r = by_path.get(str(p))
                if not r:
                    continue
                # globalny bezpiecznik: jeśli conf poniżej min_conf – nie zapisuj nic
                if conf < min_conf:
                    continue
                if conf >= hard_t and getattr(args, "set_target", False):
                    r["target_subfolder"] = f"READY TO PLAY/{bucket}"
                    r["ai_guess_bucket"] = ""
                    r["ai_guess_comment"] = f"ml; conf={conf:.2f}"
                    set_cnt += 1
                elif conf >= suggest_t and getattr(args, "suggest", False):
                    r["ai_guess_bucket"] = f"READY TO PLAY/{bucket}"
                    r["ai_guess_comment"] = f"ml; conf={conf:.2f}"
                    sug_cnt += 1
        except Exception as e:
            print(f"[WARN] ML predict failed for {p}: {e}")

    if (getattr(args, "set_target", False) or getattr(args, "suggest", False)) and (set_cnt or sug_cnt):
        save_records(CSV_PATH, rows)
    if set_cnt or sug_cnt:
        print(f"ML Predict: set={set_cnt}, suggested={sug_cnt}")


def _strip_ready_prefix(target: str) -> str:
    t = (target or "").strip()
    if t.startswith("READY TO PLAY/"):
        return t.split("/", 2)[-1] if "/" in t else t
    if t.startswith("REVIEW QUEUE/"):
        return t.split("/", 2)[-1] if "/" in t else t
    return t


def cmd_ml_train_local(args: argparse.Namespace) -> None:
    """Wytrenuj model ML na Twoich lokalnych bucketach z CSV.

    Zbiera utwory z ustawionym target_subfolder, zapewnia analizę Essentia,
    tworzy wektor cech i trenuje model. Filtruje rzadkie klasy.
    """
    if SimpleMLBucketAssigner is None:
        print("ML assigner not available (scikit-learn missing or import failure).")
        return
    if audio_analyze is None or get_analysis is None:
        print("Audio backend cache niedostępny — zainstaluj Essentia.")
        return

    rows = load_records(CSV_PATH)
    # Kandydaci: mają target_subfolder i istniejący plik (preferuj final_path)
    candidates = []
    for r in rows:
        tgt = (r.get("target_subfolder") or "").strip()
        if not tgt:
            continue
        p = None
        fp1 = r.get("final_path") or ""
        fp2 = r.get("file_path") or ""
        f1 = Path(fp1) if fp1 else None
        f2 = Path(fp2) if fp2 else None
        if f1 and f1.exists():
            p = f1
        elif f2 and f2.exists():
            p = f2
        if p is None:
            continue
        candidates.append((p, tgt))

    if not candidates:
        print("Brak danych z etykietami (target_subfolder) do treningu.")
        return

    print(f"Znaleziono {len(candidates)} oznaczonych utworów do treningu…")

    # Zbuduj labeled_tracks
    from djlib.audio.cache import compute_audio_id as _cmp
    labeled_tracks = []
    for p, tgt in candidates:
        try:
            aid = _cmp(p)
            res = get_analysis(aid)
            if (res is None) or bool(getattr(args, "recompute", False)):
                # policz analizę jeśli brak
                _ = audio_analyze(p)
                res = get_analysis(aid)
            if not res:
                continue
            feat = _normalize_ml_features_from_analysis(res)
            feat["bucket"] = _strip_ready_prefix(tgt)
            labeled_tracks.append(feat)
        except Exception:
            continue

    if not labeled_tracks:
        print("Brak gotowych cech do treningu.")
        return

    # Filtrowanie rzadkich klas
    from collections import Counter
    cnt = Counter(t["bucket"] for t in labeled_tracks)
    min_per_class = int(getattr(args, "min_per_class", 20) or 0)
    kept = [t for t in labeled_tracks if cnt[t["bucket"]] >= min_per_class]
    dropped_labels = sorted({b for b, n in cnt.items() if n < min_per_class})
    if not kept:
        print("Po odfiltrowaniu rzadkich klas nie zostało danych. Obniż --min-per-class.")
        return

    # Limit próbek
    limit = getattr(args, "limit", None)
    if isinstance(limit, int) and limit and limit > 0:
        kept = kept[:limit]

    print(f"Trenuję na {len(kept)} próbkach, klasy: {sorted({t['bucket'] for t in kept})}")
    if dropped_labels:
        print(f"Pominięte (zbyt mało próbek): {', '.join(dropped_labels)}")

    # Trening i zapis modelu
    assigner = SimpleMLBucketAssigner()
    assigner.train(kept)
    out_path = Path(getattr(args, "out", REPO_ROOT / "models" / "local_trained_model.pkl"))
    assigner.save_model(out_path)


def cmd_qa_acceptance(args: argparse.Namespace) -> None:
    """Policz acceptance rate po predykcjach ML.

    Domyślnie czyta LOGS/ml_predictions.csv i porównuje z CSV (target_subfolder).
    Zlicza, ile predykcji zostało zaakceptowanych (predykcja == docelowy bucket).
    """
    log_path = LOGS_DIR / "ml_predictions.csv"
    if not log_path.exists():
        print("Brak LOGS/ml_predictions.csv — najpierw uruchom ml-predict.")
        return
    rows = load_records(CSV_PATH)
    by_path = {r.get("file_path"): r for r in rows}

    import csv as _csv
    total = 0
    accepted = 0
    min_conf = float(getattr(args, "min_confidence", 0.65))
    per_bucket = {}
    with log_path.open("r", encoding="utf-8") as f:
        reader = _csv.DictReader(f)
        for rec in reader:
            try:
                fp = rec.get("file_path") or rec.get("file") or rec.get("path") or rec.get("0") or rec.get("file_path, bucket, confidence")
                # nasze logi mają prosty format "file_path,bucket,confidence" bez nagłówków przy pierwszym zapisie,
                # ale później dokładamy nagłówek — obsłużmy oba przypadki
                if not fp and len(rec) == 3:
                    # spróbuj odczytu z anonimowych kluczy
                    vals = list(rec.values())
                    fp, pbucket, confs = vals[0], vals[1], vals[2]
                else:
                    pbucket = rec.get("bucket") or ""
                    confs = rec.get("confidence") or "0"
                conf = 0.0
                try:
                    conf = float(confs)
                except Exception:
                    pass
                if conf < min_conf:
                    continue
                total += 1
                r = by_path.get(fp)
                if not r:
                    continue
                tgt = _strip_ready_prefix(r.get("target_subfolder", ""))
                pred = _strip_ready_prefix(pbucket)
                if tgt and pred and tgt == pred:
                    accepted += 1
                    per_bucket[pred] = per_bucket.get(pred, 0) + 1
            except Exception:
                continue

    rate = (accepted / total) if total else 0.0
    print(f"Acceptance: {accepted}/{total} = {rate:.2%}")
    if per_bucket:
        print("Akceptacje per bucket:")
        for b, c in sorted(per_bucket.items(), key=lambda x: x[1], reverse=True):
            print(f"  - {b}: {c}")


def cmd_export_xlsx(args: argparse.Namespace) -> None:
    """Eksport CSV do XLSX z dropdownem dla target_subfolder (lista bucketów).

    Edytuj w Excel/Numbers, potem użyj import-xlsx do wczytania zmian.
    """
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from typing import cast
    except Exception:
        print("Brak openpyxl — zainstaluj zależności (requirements.txt).")
        return
    from djlib.taxonomy import load_taxonomy, allowed_targets

    # Domyślnie eksportuj WSZYSTKIE rekordy; przefiltruj tylko jeśli podano --only-pending
    only_pending = bool(getattr(args, "only_pending", False))
    out_path = Path(getattr(args, "out", LOGS_DIR / "library_edit.xlsx"))

    rows = load_records(CSV_PATH)
    if only_pending:
        rows = [r for r in rows if (r.get("review_status") or "").lower() != "accepted"]

    # Wczytaj (opcjonalnie) ostatnie predykcje ML, aby pokazać audio_genre i confidence
    ml_pred_by_path: Dict[str, tuple[str, float]] = {}
    try:
        import csv as _csv
        log_path = LOGS_DIR / "ml_predictions.csv"
        if log_path.exists():
            with log_path.open("r", encoding="utf-8") as f:
                reader = _csv.DictReader(f)
                for rec in reader:
                    fp = rec.get("file_path") or rec.get("file") or rec.get("path") or rec.get("0")
                    if not fp:
                        # no headers case handled earlier when writing; skip ambiguous
                        vals = list(rec.values())
                        if len(vals) >= 3:
                            fp = vals[0]
                            pb = vals[1]
                            try:
                                cf = float(vals[2])
                            except Exception:
                                cf = 0.0
                            ml_pred_by_path[str(fp)] = (str(pb), cf)
                            continue
                    pb = rec.get("bucket") or ""
                    try:
                        cf = float(rec.get("confidence") or 0)
                    except Exception:
                        cf = 0.0
                    if fp:
                        ml_pred_by_path[str(fp)] = (pb, cf)
    except Exception:
        pass

    wb = openpyxl.Workbook()
    ws = cast("openpyxl.worksheet.worksheet.Worksheet", wb.active)  # type: ignore[name-defined]
    ws.title = "Library"  # type: ignore[assignment]
    headers = [
        "track_id", "artist", "title", "artist_suggest", "title_suggest", "version_info", "version_suggest",
        "genre", "genre_suggest",
        "bpm", "key_camelot",
        "genres_musicbrainz", "genres_lastfm", "genres_spotify", "genres_soundcloud", "audio_genre", "audio_confidence",
        "pop_playcount", "pop_listeners",
        "ai_guess_bucket", "ai_guess_comment", "target_subfolder", "file_path"
    ]
    ws.append(headers)
    for r in rows:
        fp = r.get("file_path", "")
        audio_genre = ""
        audio_conf = ""
        if fp and fp in ml_pred_by_path:
            pb, cf = ml_pred_by_path[fp]
            audio_genre = pb
            audio_conf = f"{cf:.2f}"
        # Round BPM to integer (display only)
        bpm_raw = (r.get("bpm") or "").strip()
        bpm_disp = ""
        try:
            if bpm_raw:
                bpm_disp = str(int(round(float(bpm_raw))))
        except Exception:
            bpm_disp = bpm_raw
        ws.append([
            r.get("track_id", ""),
            r.get("artist", ""),
            r.get("title", ""),
            r.get("artist_suggest", ""),
            r.get("title_suggest", ""),
            r.get("version_info", ""),
            r.get("version_suggest", ""),
            r.get("genre", ""),
            r.get("genre_suggest", ""),
            bpm_disp,
            r.get("key_camelot", ""),
            r.get("genres_musicbrainz", ""),
            r.get("genres_lastfm", ""),
            r.get("genres_spotify", ""),
            r.get("genres_soundcloud", ""),
            audio_genre,
            audio_conf,
            r.get("pop_playcount", ""),
            r.get("pop_listeners", ""),
            r.get("ai_guess_bucket", ""),
            r.get("ai_guess_comment", ""),
            r.get("target_subfolder", ""),
            fp,
        ])

    # Arkusz z bucketami
    # Użyj pełnych nazw z prefiksami (READY TO PLAY/…, REVIEW QUEUE/…)
    tax = load_taxonomy()
    buckets = allowed_targets()
    ws2 = wb.create_sheet("Buckets")
    for i, b in enumerate(buckets, start=1):
        ws2.cell(row=i, column=1, value=b)
    # Data validation: dopasuj do kolumny 'target_subfolder' niezależnie od zmian układu kolumn
    max_row = ws.max_row
    dv = DataValidation(type="list", formula1=f"=Buckets!$A$1:$A${len(buckets)}", allow_blank=True)
    try:
        tgt_idx = headers.index("target_subfolder") + 1
        col_letter = get_column_letter(tgt_idx)
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)  # type: ignore[arg-type]
    except ValueError:
        # jeśli brak kolumny (nie powinno się zdarzyć), pomiń walidację, ale zapisz plik
        pass
    ws.add_data_validation(dv)
    # kosmetyka
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Zapisano: {out_path}")


def cmd_import_xlsx(args: argparse.Namespace) -> None:
    """Wczytaj zmiany z XLSX (kolumna target_subfolder) z powrotem do CSV.

    Dopasowanie po track_id; jeśli brak, próba po file_path.
    """
    try:
        import openpyxl
    except Exception:
        print("Brak openpyxl — zainstaluj zależności (requirements.txt).")
        return

    src = Path(getattr(args, "path", LOGS_DIR / "library_edit.xlsx"))
    if not src.exists():
        print(f"Brak pliku: {src}")
        return
    wb = openpyxl.load_workbook(src)
    ws = wb["Library"]

    rows = load_records(CSV_PATH)
    by_id = {r.get("track_id"): r for r in rows}
    by_path = {r.get("file_path"): r for r in rows}

    # Map header indices
    header = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    def _cell(row_vals, key):
        i = idx.get(key)
        return row_vals[i] if i is not None and i < len(row_vals) else None

    updated = 0
    accepted = 0
    for i, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        track_id = str(_cell(row_vals, "track_id") or "").strip()
        file_path = str(_cell(row_vals, "file_path") or "").strip()
        rec = by_id.get(track_id) or by_path.get(file_path)
        if not rec:
            continue
        # Target subfolder
        tgt = str(_cell(row_vals, "target_subfolder") or "").strip()
        if tgt and tgt != rec.get("target_subfolder", ""):
            rec["target_subfolder"] = tgt
            updated += 1
        # Accept metadata edits: if user filled artist/title/genre columns
        artist_edit = str(_cell(row_vals, "artist") or "").strip()
        title_edit = str(_cell(row_vals, "title") or "").strip()
        version_info_edit = str(_cell(row_vals, "version_info") or "").strip()
        genre_edit = str(_cell(row_vals, "genre") or "").strip()
        if artist_edit and artist_edit != rec.get("artist", ""):
            rec["artist"] = artist_edit
            updated += 1
        if title_edit and title_edit != rec.get("title", ""):
            rec["title"] = title_edit
            updated += 1
        if version_info_edit and version_info_edit != rec.get("version_info", ""):
            rec["version_info"] = version_info_edit
            updated += 1
        if genre_edit and genre_edit != rec.get("genre", ""):
            rec["genre"] = genre_edit
            updated += 1
        # If user modified any accepted fields or set target, mark as accepted
        if tgt or artist_edit or title_edit or genre_edit:
            if (rec.get("review_status") or "") != "accepted":
                rec["review_status"] = "accepted"
                accepted += 1
        # Sync suggests to accepted if accepted but main empty
        if rec.get("review_status") == "accepted":
            if not rec.get("artist") and rec.get("artist_suggest"):
                rec["artist"] = rec["artist_suggest"]
            if not rec.get("title") and rec.get("title_suggest"):
                rec["title"] = rec["title_suggest"]
            if not rec.get("version_info") and rec.get("version_suggest"):
                rec["version_info"] = rec["version_suggest"]
            if not rec.get("genre") and rec.get("genre_suggest"):
                rec["genre"] = rec["genre_suggest"].split(",")[0].strip()

    if updated or accepted:
        save_records(CSV_PATH, rows)
    print(f"Import XLSX: updated={updated}, accepted={accepted}")


# ============ META-KOMENDY (aliasy) ============

def cmd_round_1(args: argparse.Namespace) -> None:
    """Uruchom pełny pierwszy etap: analyze → enrich-online → ml-predict (suggest) → export-xlsx."""
    # analyze-audio
    aargs = argparse.Namespace(
        path=str(INBOX_DIR), check_env=False, recompute=False, workers=1, target_bpm="80:180"
    )
    cmd_analyze_audio(aargs)

    # enrich-online
    cmd_enrich_online(argparse.Namespace())

    # ml-predict (sugestie)
    pargs = argparse.Namespace(
        model=str(REPO_ROOT / "models" / "fma_trained_model_balanced.pkl"),
        path=str(INBOX_DIR),
        recompute=False,
        set_target=False,
        suggest=True,
        hard_threshold=0.85,
        suggest_threshold=float(getattr(args, "suggest_threshold", 0.65)),
        min_confidence=float(getattr(args, "min_confidence", 0.40)),
    )
    cmd_ml_predict(pargs)

    # export-xlsx
    exargs = argparse.Namespace(
        out=str(getattr(args, "xlsx_out", LOGS_DIR / "library_edit.xlsx")),
        only_pending=True,
    )
    cmd_export_xlsx(exargs)
    print(f"➡️  Otwórz do edycji: {exargs.out}")


def cmd_round_2(args: argparse.Namespace) -> None:
    """Uruchom drugi etap: import-xlsx → apply → ml-train-local → qa-acceptance."""
    # import-xlsx
    imargs = argparse.Namespace(path=str(getattr(args, "xlsx_path", LOGS_DIR / "library_edit.xlsx")))
    cmd_import_xlsx(imargs)

    # apply
    cmd_apply(argparse.Namespace(dry_run=False))

    # ml-train-local
    tlargs = argparse.Namespace(
        min_per_class=int(getattr(args, "min_per_class", 20)),
        limit=None,
        recompute=False,
        out=str(getattr(args, "model_out", REPO_ROOT / "models" / "local_trained_model.pkl"))
    )
    cmd_ml_train_local(tlargs)

    # qa-acceptance
    qaargs = argparse.Namespace(min_confidence=float(getattr(args, "qa_min_confidence", 0.65)))
    cmd_qa_acceptance(qaargs)

# ============ PARSER ============

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="djlib", description="DJ Library Manager CLI")
    sp = p.add_subparsers(dest="cmd", required=True)

    sp.add_parser("configure").set_defaults(func=cmd_configure)
    sp.add_parser("scan").set_defaults(func=cmd_scan)

    ap = sp.add_parser("auto-decide")
    ap.add_argument("--rules", default=str(REPO_ROOT / "rules.yml"))
    ap.add_argument("--only-empty", action="store_true")
    ap.set_defaults(func=cmd_auto_decide)

    ap2 = sp.add_parser("apply")
    ap2.add_argument("--dry-run", action="store_true")
    ap2.set_defaults(func=cmd_apply)

    sp.add_parser("undo").set_defaults(func=cmd_undo)
    sp.add_parser("dupes").set_defaults(func=cmd_dupes)
    sap = sp.add_parser("sync-audio-metrics")
    sap.add_argument("--force", action="store_true")
    sap.add_argument("--write-tags", action="store_true", help="Zapisz metadane (BPM/Key) do plików audio")
    sap.set_defaults(func=cmd_sync_audio_metrics)
    sp.add_parser("fix-fingerprints").set_defaults(func=cmd_fix_fingerprints)
    ep = sp.add_parser("enrich-online")
    ep.add_argument("--force-genres", action="store_true", help="Nadpisz kolumny genres_musicbrainz/lastfm/spotify nawet jeśli już wypełnione")
    ep.add_argument("--skip-soundcloud", action="store_true", help="Pomiń źródło SoundCloud nawet jeśli client_id jest ustawiony")
    ep.set_defaults(func=cmd_enrich_online)

    # analyze-audio
    aap = sp.add_parser("analyze-audio")
    aap.add_argument("--path", default=str(INBOX_DIR), help="Ścieżka pliku lub folderu (domyślnie INBOX)")
    aap.add_argument("--check-env", action="store_true", help="Sprawdź środowisko Essentia")
    aap.add_argument("--recompute", action="store_true", help="Pomiń cache i przelicz na nowo")
    aap.add_argument("--workers", type=int, default=1, help="Liczba workerów (na razie ignorowane; skeleton)")
    aap.add_argument("--target-bpm", default="80:180", help="Zakres docelowy BPM, np. 80:180")
    aap.set_defaults(func=cmd_analyze_audio)

    # ml predict
    mp = sp.add_parser("ml-predict")
    mp.add_argument("--model", default=str(REPO_ROOT / "models" / "fma_trained_model_balanced.pkl"))
    mp.add_argument("--path", default=str(INBOX_DIR), help="Plik lub folder (domyślnie INBOX)")
    mp.add_argument("--recompute", action="store_true", help="Przelicz analizę audio na nowo")
    mp.add_argument("--set-target", action="store_true", help="Ustawiaj docelowy kubełek powyżej progu hard")
    mp.add_argument("--suggest", action="store_true", help="Ustawiaj tylko ai_guess_* powyżej progu suggest")
    mp.add_argument("--hard-threshold", type=float, default=0.85)
    mp.add_argument("--suggest-threshold", type=float, default=0.65)
    mp.add_argument("--min-confidence", type=float, default=0.40, help="Nie zapisuj żadnych sugestii poniżej tego progu")
    mp.set_defaults(func=cmd_ml_predict)

    # ml-train-local: trenuj model na Twoich zaakceptowanych bucketach
    tl = sp.add_parser("ml-train-local")
    tl.add_argument("--min-per-class", type=int, default=20, help="Minimalna liczba próbek na klasę (odfiltruj rzadkie)")
    tl.add_argument("--limit", type=int, default=None, help="Limit próbek do szybkiego treningu (opcjonalnie)")
    tl.add_argument("--recompute", action="store_true", help="Przelicz analizę Essentia jeśli brak w cache")
    tl.add_argument("--out", default=str(REPO_ROOT / "models" / "local_trained_model.pkl"))
    tl.set_defaults(func=cmd_ml_train_local)

    # QA: acceptance rate
    qa = sp.add_parser("qa-acceptance")
    qa.add_argument("--min-confidence", type=float, default=0.65, help="Licz tylko predykcje powyżej tego progu")
    qa.set_defaults(func=cmd_qa_acceptance)

    # XLSX export/import z dropdownem na bucket
    ex = sp.add_parser("export-xlsx")
    ex.add_argument("--out", default=str(LOGS_DIR / "library_edit.xlsx"))
    ex.add_argument("--only-pending", action="store_true", help="Eksportuj tylko niezaakceptowane")
    ex.set_defaults(func=cmd_export_xlsx)

    im = sp.add_parser("import-xlsx")
    im.add_argument("--path", default=str(LOGS_DIR / "library_edit.xlsx"))
    im.set_defaults(func=cmd_import_xlsx)

    # genres resolve (single lookup)
    gp = sp.add_parser("genres")
    gsp = gp.add_subparsers(dest="subcmd", required=True)
    res = gsp.add_parser("resolve")
    res.add_argument("--artist", required=True)
    res.add_argument("--title", required=True)
    res.add_argument("--duration", type=int, default=None, help="Duration in seconds (optional)")
    res.set_defaults(func=cmd_genres_resolve)

    sp.add_parser("detect-taxonomy").set_defaults(func=cmd_detect_taxonomy)

    # --- Meta-komendy: round-1 i round-2 ---
    r1 = sp.add_parser("round-1", help="Analyze + Enrich + ML Predict (suggest) + Export XLSX")
    r1.add_argument("--min-confidence", type=float, default=0.40)
    r1.add_argument("--suggest-threshold", type=float, default=0.65)
    r1.add_argument("--xlsx-out", default=str(LOGS_DIR / "library_edit.xlsx"))
    r1.set_defaults(func=cmd_round_1)

    r2 = sp.add_parser("round-2", help="Import XLSX + Apply + Train local model + QA acceptance")
    r2.add_argument("--xlsx-path", default=str(LOGS_DIR / "library_edit.xlsx"))
    r2.add_argument("--min-per-class", type=int, default=20)
    r2.add_argument("--qa-min-confidence", type=float, default=0.65)
    r2.add_argument("--model-out", default=str(REPO_ROOT / "models" / "local_trained_model.pkl"))
    r2.set_defaults(func=cmd_round_2)

    tb = sp.add_parser("taxonomy-backup", help="Zrób snapshot taksonomii na podstawie folderów i zapisz backup")
    tb.set_defaults(func=cmd_taxonomy_backup)
    return p

def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
