from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Mapping, Sequence
import csv

from djlib.filename import build_final_filename
from djlib.locks import csv_lock


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    width: float | None = None
    hidden: bool = False
    locked: bool = False  # True = read-only (Review UI uses this for display)


UNSORTED_COLUMNS: Sequence[ColumnSpec] = [
    # Read-only metadata (locked)
    ColumnSpec("track_id", hidden=True, width=22, locked=True),
    ColumnSpec("rekordbox_id", hidden=True, width=16, locked=True),
    ColumnSpec("traktor_id", hidden=True, width=16, locked=True),
    ColumnSpec("file_path", width=115, locked=True),  # Visible for reference
    ColumnSpec("file_hash", hidden=True, width=30, locked=True),
    ColumnSpec("fingerprint", hidden=True, width=26, locked=True),
    ColumnSpec("added_date", hidden=True, width=18, locked=True),
    ColumnSpec("is_duplicate", hidden=True, width=12, locked=True),
    ColumnSpec("duplicate_paths", hidden=True, width=60, locked=True),
    ColumnSpec("tag_artist_original", width=26, locked=True),  # Visible
    ColumnSpec("tag_title_original", width=26, locked=True),   # Visible
    ColumnSpec("tag_genre_original", width=22, locked=True),   # Visible
    ColumnSpec("tag_bpm_original", hidden=True, width=14, locked=True),
    ColumnSpec("tag_key_original", hidden=True, width=14, locked=True),
    ColumnSpec("artist_suggest", hidden=True, width=24, locked=True),
    ColumnSpec("title_suggest", hidden=True, width=24, locked=True),
    ColumnSpec("title_normalized", hidden=False, width=12, locked=True),  # "yes" if title was normalized from MB canonical
    ColumnSpec("version_suggest", hidden=True, width=20, locked=True),
    ColumnSpec("genre_suggest", hidden=False, width=24, locked=True),
    ColumnSpec("album_suggest", hidden=True, width=32, locked=True),  # Hidden - user manages own albums
    ColumnSpec("release_group_id", hidden=True, width=36, locked=True),  # MusicBrainz MBID (kept for reference)
    ColumnSpec("year_suggest", hidden=False, width=12, locked=True),
    ColumnSpec("duration_suggest", hidden=True, width=16, locked=True),
    # Genre hints (visible for decision making, but locked)
    ColumnSpec("genres_musicbrainz", width=24, locked=True),
    ColumnSpec("genres_lastfm", width=24, locked=True),
    ColumnSpec("genres_soundcloud", width=24, locked=True),
    ColumnSpec("genres_beatport", width=24, locked=True),
    ColumnSpec("pop_playcount", width=14, locked=True),
    ColumnSpec("pop_listeners", width=14, locked=True),
    ColumnSpec("meta_source", hidden=True, width=20, locked=True),
    # MusicBrainz canonical first release data (hidden, locked)
    ColumnSpec("original_album_title", hidden=True, width=32, locked=True),
    ColumnSpec("original_release_date", hidden=True, width=18, locked=True),
    ColumnSpec("original_release_year", hidden=True, width=12, locked=True),
    ColumnSpec("original_release_mbid", hidden=True, width=36, locked=True),
    ColumnSpec("original_release_group_mbid", hidden=True, width=36, locked=True),
    ColumnSpec("original_release_category", hidden=True, width=16, locked=True),
    ColumnSpec("original_release_source", hidden=True, width=24, locked=True),
    # Editable fields (user reviews/accepts via Review UI)
    ColumnSpec("artist", width=30),
    ColumnSpec("title", width=45),
    ColumnSpec("version_info", width=42),
    ColumnSpec("year", width=10),
    ColumnSpec("genre", width=24),  # User-selected genre (dropdown from genres.yml)
    ColumnSpec("genre_mapping_status", width=18, locked=True),  # "OK" or "UNMAPPED"
    ColumnSpec("status", width=12),  # accept / reject / review
    ColumnSpec("destination", width=14),  # library / reject / archive
    ColumnSpec("must_play", width=14),
    ColumnSpec("occasion_tags", width=24),
    ColumnSpec("notes", width=40),
    ColumnSpec("bpm", width=10),
    ColumnSpec("key_camelot", width=12),
    ColumnSpec("rating", width=10),  # Star rating 0-5 (syncs with Rekordbox/Traktor)
    # Computed final filename (locked)
    ColumnSpec("final_filename", width=100, locked=True),
    # AI classification results (locked — set by batch_classify / api)
    ColumnSpec("ai_artist", hidden=True, width=30, locked=True),
    ColumnSpec("ai_title", hidden=True, width=45, locked=True),
    ColumnSpec("ai_version", hidden=True, width=42, locked=True),
    ColumnSpec("ai_genre", hidden=True, width=24, locked=True),
    ColumnSpec("ai_confidence", hidden=True, width=10, locked=True),
    ColumnSpec("ai_reasoning", hidden=True, width=60, locked=True),
    ColumnSpec("ai_classify_date", hidden=True, width=18, locked=True),
    # Status column (editable)
    ColumnSpec("done", width=10),
]

# Canonical fieldnames list for CSV I/O
UNSORTED_FIELDNAMES: List[str] = [col.name for col in UNSORTED_COLUMNS]

DONE_CHOICES = ("TRUE", "FALSE")
STATUS_CHOICES = ("accept", "reject", "review", "")
DESTINATION_CHOICES = ("library", "reject", "archive", "mixes", "")


def _as_str(val: object | None) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    return str(val)


def normalize_unsorted_row(row: Mapping[str, str | None]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for col in UNSORTED_COLUMNS:
        out[col.name] = _as_str(row.get(col.name, ""))
    if not out.get("done"):
        out["done"] = "FALSE"
    
    # Compute final_filename preview
    artist = out.get("artist", "")
    title = out.get("title", "")
    version = out.get("version_info", "")
    bpm = out.get("bpm", "")
    key = out.get("key_camelot", "")
    # Get extension from file_path if available
    file_path = out.get("file_path", "")
    ext = Path(file_path).suffix if file_path else ".mp3"
    
    out["final_filename"] = build_final_filename(artist, title, version, key, bpm, ext)
    
    return out


# ── XLSX → CSV migration ────────────────────────────────────────────────────

def _migrate_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> bool:
    """One-time migration: convert existing unsorted.xlsx to unsorted.csv.

    Returns True if migration happened, False otherwise.
    """
    if not xlsx_path.exists():
        return False
    try:
        from openpyxl import load_workbook
    except ImportError:
        # openpyxl not available — can't migrate
        print(f"⚠️  Found {xlsx_path.name} but openpyxl not installed for migration.")
        print(f"   Install openpyxl temporarily: pip install openpyxl")
        return False

    print(f"📦 Migrating {xlsx_path.name} → {csv_path.name} …")
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [_as_str(cell.value) for cell in header_row]
    rows: List[Dict[str, str]] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row:
            continue
        if all(v in (None, "") for v in excel_row):
            continue
        rec: Dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = excel_row[idx] if idx < len(excel_row) else ""
            rec[header] = _as_str(value)
        rows.append(normalize_unsorted_row(rec))

    # Write CSV
    write_unsorted_rows(csv_path, rows)
    print(f"✅ Migrated {len(rows)} rows to {csv_path.name}")

    # Rename old xlsx to .xlsx.bak
    bak = xlsx_path.with_suffix(".xlsx.bak")
    xlsx_path.rename(bak)
    print(f"   Old file backed up as {bak.name}")
    return True


# ── CSV I/O ──────────────────────────────────────────────────────────────────

def load_unsorted_rows(path: Path) -> List[Dict[str, str]]:
    """Load unsorted rows from CSV file.

    If path doesn't exist but a corresponding .xlsx does, auto-migrates first.
    """
    # Auto-migration: check for legacy .xlsx alongside expected .csv
    if not path.exists() and path.suffix == ".csv":
        xlsx_path = path.with_suffix(".xlsx")
        if xlsx_path.exists():
            _migrate_xlsx_to_csv(xlsx_path, path)

    if not path.exists():
        return []

    rows: List[Dict[str, str]] = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for rec in reader:
            rows.append(normalize_unsorted_row(rec))
    return rows


def write_unsorted_rows(
    path: Path,
    rows: Iterable[Dict[str, str]],
    bucket_choices: Sequence[str] | None = None,
) -> None:
    """Write rows to unsorted CSV.

    Args:
        path: Path to unsorted.csv
        rows: Row data dictionaries
        bucket_choices: Legacy parameter (ignored, kept for API compatibility)
    """
    normalized_rows = [normalize_unsorted_row(r) for r in rows]

    path.parent.mkdir(parents=True, exist_ok=True)
    # Lock to serialize against concurrent Review UI saves and CLI runs.
    # Re-entrant: callers doing read-modify-write can hold the lock externally.
    with csv_lock(path):
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=UNSORTED_FIELDNAMES, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(normalized_rows)


def is_done(value: str | None) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in {"TRUE", "YES", "1", "DONE"}
