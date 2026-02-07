#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('data/unsorted.xlsx')
ws = wb.active

# Get header row to find column indices
headers = [cell.value for cell in ws[1]]

# Find column indices
cols = {h: i for i, h in enumerate(headers)}

# Print tracks without genre
print('=== TRACKI BEZ GATUNKU ===')
count = 0
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    genre = row[cols.get('genre')] if 'genre' in cols else None
    if not genre or genre == 'None':
        artist = row[cols.get('artist', 0)] or row[cols.get('artist_suggest', 0)] or ''
        title = row[cols.get('title', 0)] or row[cols.get('title_suggest', 0)] or ''
        version_info = row[cols.get('version_info', 0)] or ''
        version_suggest = row[cols.get('version_suggest', 0)] or ''
        print(f'{idx}: {artist} - {title}')
        print(f'    version_info={version_info!r}')
        print(f'    version_suggest={version_suggest!r}')
        count += 1

print(f'\n=== RAZEM: {count} tracków bez gatunku ===')
